import rich_click as click
from pathlib import Path
import pandas as pd
from loguru import logger
import sys
import numpy as np
import shutil
from openpyxl import load_workbook
from datetime import datetime
import re

# Configure logger
logger.remove()  # Remove default handler
logger.add(sys.stderr, level="INFO")  # Add handler with stderr as sink

# Add this near the top after the rich_click import
click.rich_click.USE_RICH_MARKUP = True

FIRST_YEAR = 2025

TEMPLATE_PATH = (
    Path(__file__).parent.parent.parent
    / "data"
    / "msm22-output-templates"
    / "FE_template.xlsx"
)

CSV_COLUMN_ORDER_MAPPING = {
    "Elec fuels": ["model", "study", "region", "isp_subregion", "year", "unit", "varbl", "fuel", "scen", "tech", "val"],
    "CO2 emissions - non bldg+ind": ["model", "region", "isp_subregion", "year", "unit", "emission_type", "enduse", "scen", "sector", "state", "subsector_p", "tech", "val"],
    "CO2 emissions - Industry - Process": ["region", "isp_subregion", "year", "unit", "ee_category", "endusegroup_p", "scen", "source", "val"],
    "Elec capacity and generation": ["model", "study", "region", "isp_subregion", "year", "unit", "varbl", "fuel", "scen", "tech", "val"],
    "EnEff Buildings": ["model", "study", "region", "isp_subregion","year", "unit", "varbl", "buildingtype", "ee_category", "enduse", "fuel", "scen", "source", "val"],
    "EnEff Industry": ["model", "study", "region", "isp_subregion", "year", "unit", "varbl", "ee_category", "fuel", "nemreg", "scen", "source", "subsectorgroup_c", "val"],
    "Fin Energy Residential": ["model", "study", "region", "isp_subregion", "year", "unit", "varbl", "enduse", "fuel", "fuel_switched", "scen", "source", "subsector_p", "val"],
    "Fin energy Transport": ["region", "isp_subregion","year", "unit", "varbl", "enduse", "fuel", "scen", "subsector_p", "tech", "val"],
    "Fuels switched industry": ["sector", "scen", "region", "isp_subregion", "year", "source", "subsectorgroup_c", "fuel_switched_from", "fuel_switched_to", "hydrogen_source", "PJ_switched"],
    "Hydrogen capacity and generation": ["model", "study", "region", "year", "unit", "varbl", "process", "scen", "tech", "val"],
    "Hydrogen exports": ["model", "study", "region", "year", "unit", "varbl", "scen", "val"],
    "Hydrogen fuels": ["model", "study", "region", "year", "unit", "varbl", "process", "commodity", "fuel", "scen", "tech", "val"],
    # These two match the columns in the FE_template.xlsx, not the final CSVs
    "Commercial FE with EnInt": ["sector_p", "scen", "region", "isp_subregion", "year", "enduse", "source_p", "buildingtype", "fuel", "fuel_override", "fuel_switched", "IESTCS_EnInt", "IESTCS_Out"],
    "Industry FE with EnInt": ["sector_p", "scen", "region", "isp_subregion", "year", "process", "source_p", "hydrogen_source", "fuel", "fuel_override", "subsectorgroup_c", "IESTCS_EnInt", "IESTCS_Out"],
}


CSV_TO_FILTER_OUT_MAPPING = {
    "CO2 emissions - Industry - Process": [{"source":["-"]}],
}


# Add this mapping dictionary before the main function
SHEET_NAME_MAPPING = {
    "MSM22 CO2 emis-ind-proc": "CO2 emissions - Industry - Process",
    "MSM22 emis-non-bldg+ind": "CO2 emissions - non bldg+ind",
    "MSM22 Commercial FE with EnInt": "Fin Energy Commercial",
    "MSM22 Industry FE with EnInt": "Fin Energy Industry",
    "MSM22 Elec cap and gen": "Elec capacity and generation",
    "MSM22 Elec fuels": "Elec fuels",
    "MSM22 EnEff Buildings": "EnEff Buildings",
    "MSM22 EnEff Industry": "EnEff Industry",
    "MSM22 Fin energy res": "Fin Energy Residential",
    "MSM22 Fin energy tra": "Fin energy Transport",
    "MSM22 Industry FE with EnInt": "Fuels switched industry",
    "MSM22 h2-cap-and-gen": "Hydrogen capacity and generation",
    "MSM22 Hydrogen exports": "Hydrogen exports",
    "MSM22 Hydrogen fuels": "Hydrogen fuels",
    "MSM22-ind2-emissions": "IND2 emissions",
    "MSM22 Commercial FE with EnInt": "Commercial FE with EnInt",
    "MSM22 Industry FE with EnInt": "Industry FE with EnInt",
}

# Add this near the top of the file with other imports
pd.set_option("future.no_silent_downcasting", True)

def is_valid_results_file(filename: str) -> bool:
    """Check if filename matches the results_YYYYMMDD-HHMMSS.xlsx pattern."""
    pattern = r"results_\d{8}-\d{6}\.xlsx$"
    return bool(re.match(pattern, filename))

def extract_datetime(filename: str) -> datetime:
    """Extract datetime from results filename."""
    # Extract YYYYMMDD-HHMMSS from the filename
    date_str = re.search(r"(\d{8}-\d{6})", filename).group(1)
    return datetime.strptime(date_str, "%Y%m%d-%H%M%S")

def process_msm22_csvs(input_dir: Path | str) -> None:
    """Process MSM22 files from a directory and create CSVs.

    Args:
        input_dir: Path to directory containing MSM22 Excel files
    """
    input_path = Path(input_dir).resolve()
    output_path = input_path

    logger.info(f"Reading Excel files from: {input_path}")

    # Initialize DataFrames to store combined data for each sheet type
    combined_data = {}

    # Find all Excel files in directory that match the pattern
    excel_files = [
        f for f in input_path.glob("*.xlsx") 
        if is_valid_results_file(f.name)
    ]
    
    if not excel_files:
        logger.info(f"Files found in directory: {[f.name for f in input_path.glob('*.xlsx')]}")
        raise ValueError(f"No valid results files found in {input_path}")

    # Sort files by datetime (oldest first)
    excel_files.sort(key=lambda x: extract_datetime(x.name))
    
    logger.info(f"Found {len(excel_files)} valid results files")
    for file in excel_files:
        logger.info(f"  {file.name} ({extract_datetime(file.name)})")

    # Process each Excel file
    for excel_path in excel_files:
        logger.info(f"Processing file: {excel_path}")
        excel_file = pd.ExcelFile(excel_path)

        for sheet_name in excel_file.sheet_names:
            if sheet_name == "Info":
                logger.info(f"Skipping Info sheet")
                continue

            logger.info(f"Processing sheet: {sheet_name}")
            df = pd.read_excel(excel_file, sheet_name=sheet_name, skiprows=1)

            df_raw = df.copy(deep=True)

            # Convert year column to numeric, coercing errors to NaN
            df["year"] = pd.to_numeric(df["year"], errors="coerce")

            # Drop any rows where year is < FIRST_YEAR
            df = df[df["year"] >= FIRST_YEAR]

            # Replace "-" with NaN
            df = df.replace("-", np.nan)

            # Override fuel
            if sheet_name not in ["MSM22 Commercial FE with EnInt", "MSM22 Industry FE with EnInt"]:
                if "fuel_override" in df.columns:
                    mask = df["fuel_override"].notna()
                    df.loc[mask, "fuel"] = df.loc[mask, "fuel_override"]
                    df = df.drop(columns=["fuel_override"])

            # Override isp_subregion
            if "isp_subregion_override" in df.columns and "isp_subregion" in df.columns:
                mask = df["isp_subregion_override"].notna()
                df.loc[mask, "isp_subregion"] = df.loc[mask, "isp_subregion_override"]
                df = df.drop(columns=["isp_subregion_override"])

            # Rename columns
            df = df.rename(
                columns={"GrandTotal": "val"}
            )
            if sheet_name not in ["MSM22 Commercial FE with EnInt", "MSM22 Industry FE with EnInt"]:
                df = df.rename(
                    columns={"source_p": "source", "sector_p": "sector", "GrandTotal": "val"}
                )

            # Replace NaN with "-"
            df = df.fillna("-")

            if df.empty:
                logger.error(f"Empty DataFrame for sheet '{sheet_name}', skipping")
                continue

            # Get the CSV name and column order
            csv_name = SHEET_NAME_MAPPING.get(sheet_name, None)
            if csv_name is None:
                logger.warning(f"No CSV name found for {sheet_name}, skipping")
                continue

            column_order = CSV_COLUMN_ORDER_MAPPING.get(csv_name, None)
            if column_order is None:
                logger.info(f"No column order found for {csv_name}, skipping")
                continue

            # Check columns
            if not all(col in df.columns for col in column_order):
                logger.error(f"Column mismatch in {sheet_name}:")
                logger.error(f"Found columns: {sorted(list(df.columns))}")
                logger.error(f"Expected columns: {sorted(column_order)}")
                # list the missing columns
                missing_cols = [col for col in column_order if col not in df.columns]
                logger.error(f"Missing columns: {sorted(missing_cols)}")
                continue
            # Reorder columns
            df = df[column_order]

            # Add to combined data - newer files will overwrite older ones
            csv_name = SHEET_NAME_MAPPING.get(sheet_name)

            # Group by all columns except val and sum
            if csv_name not in ["Commercial FE with EnInt", "Industry FE with EnInt"]:
                df = df.groupby(column_order[:-1], as_index=False).sum()

            # Check for nans
            if df.isna().any().any():
                logger.error(f"NaN values found in {csv_name}, skipping CSV creation")
                continue

            # Filter out rows
            if csv_name in CSV_TO_FILTER_OUT_MAPPING:
                for filter_dict in CSV_TO_FILTER_OUT_MAPPING[csv_name]:
                    logger.info(f"Filtering out rows with {filter_dict}")
                    # Log unique values in each filter column
                    for col, values in filter_dict.items():
                        unique_vals = df[col].unique()
                        logger.info(f"Unique values in {col}: {unique_vals}")
                        # Remove entire rows where this column matches any of the filtered values
                        rows_before = len(df)
                        df = df[~df[col].isin(values)]
                        rows_removed = rows_before - len(df)
                        logger.info(f"Filtered out {rows_removed} rows where {col} matched {values}")
                        # log new unique values
                        new_unique_vals = df[col].unique()
                        logger.info(f"New unique values in {col}: {new_unique_vals}")

            csv_path = output_path / f"{csv_name}.csv"
            df.to_csv(csv_path, index=False)
            logger.info(f"Created CSV file: {csv_path}")

def process_energy_intensity(input_dir: Path | str) -> None:
    """Process energy intensity files using the already processed CSVs.

    Args:
        input_dir: Path to directory containing the processed CSVs
    """
    input_path = Path(input_dir).resolve()
    output_path = input_path

    logger.info("Starting energy intensity file processing...")

    # Check if required CSVs exist
    required_csvs = ["Commercial FE with EnInt.csv", "Industry FE with EnInt.csv"]
    for csv_file in required_csvs:
        if not (input_path / csv_file).exists():
            raise ValueError(f"Required CSV file not found: {csv_file}")

    if not TEMPLATE_PATH.exists():
        raise ValueError(f"Template file not found at: {TEMPLATE_PATH}")

    # Copy template file to destination
    output_file = output_path / f"FE_processed_results.xlsx"
    shutil.copy2(TEMPLATE_PATH, Path(output_file).resolve())

    # Read the processed CSVs
    processed_dfs = {}
    template_sheets = ["Commercial FE with EnInt", "Industry FE with EnInt"]

    for sheet in template_sheets:
        csv_path = input_path / f"{sheet}.csv"
        logger.info(f"Reading processed CSV: {csv_path}")
        df = pd.read_csv(csv_path)
        processed_dfs[sheet] = df

    # Use openpyxl to update only the data cells while preserving formulas
    wb = load_workbook(output_file)

    # Before pasting, check that the column names match
    for sheet, processed_df in processed_dfs.items():
        ws = wb[sheet]
        template_columns = [col.value for col in ws[1]]
        processed_columns = processed_df.columns.tolist()
        template_columns = template_columns[:len(processed_columns)]
        if template_columns != processed_columns:
            print(f"Template columns: {template_columns}")
            print(f"CSV columns: {processed_columns}")
            raise ValueError(f"Column names do not match for sheet {sheet}")

    # Update the template with the CSV data
    for sheet, processed_df in processed_dfs.items():
        ws = wb[sheet]

        # Convert DataFrame to values and update cells directly
        data_values = processed_df.values
        for i, row in enumerate(data_values):
            for j, value in enumerate(row):
                if pd.notna(value):  # Only update non-NA values
                    ws.cell(row=i + 2, column=j + 1, value=value)

        # Handle template row count
        max_row = ws.max_row
        if max_row > len(data_values) + 1:
            rows_to_delete = max_row - (len(data_values) + 1)
            logger.info(f"Deleting {rows_to_delete} rows from the bottom")
            ws.delete_rows(len(data_values) + 2, rows_to_delete)
        elif max_row < len(data_values) + 1:
            raise ValueError(
                f"Data is longer than template by {len(data_values) + 1 - max_row} rows. Please copy formulas down in columns M:P."
            )

    wb.save(output_file)

    # Recalculate formulas using xlwings
    try:
        import xlwings as xw
        wb = xw.Book(output_file)
        wb.save()
        wb.close()
    except Exception as e:
        logger.error(f"Failed to recalculate formulas - RUN IN A TERMINAL: {e}")

    logger.info(f"Created processed energy intensity file: {output_file}")

    # Create emissions and final energy CSVs from the processed file
    for sheet in template_sheets:
        df = pd.read_excel(
            output_file,
            sheet_name=sheet,
            usecols="A:Q",
            engine_kwargs={"data_only": True},
        )
        
        # Drop rows where IESTCS_Out is NaN or missing
        df = df.dropna(subset=["IESTCS_Out"])
        logger.info(f"Columns in processed file: {df.columns.tolist()}")

        # Get the correct column order from CSV_COLUMN_ORDER_MAPPING
        cols = CSV_COLUMN_ORDER_MAPPING[sheet]
        if sheet == "Commercial FE with EnInt":
            emissions_csv_name = "CO2 emissions Commercial"
            fin_energy_csv_name = "Final Energy Commercial"
        else:
            emissions_csv_name = "CO2 emissions Industry"
            fin_energy_csv_name = "Final Energy Industry"

        # Create emissions CSV
        logger.info(f"Creating emissions CSV: {emissions_csv_name}")
        emissions_df = df[cols + ["kt"]]
        emissions_df = emissions_df.groupby(cols, as_index=False).sum()
        if not emissions_df.empty:
            emissions_df.to_csv(output_path / f"{emissions_csv_name}.csv", index=False)

        # Create final energy CSV
        logger.info(f"Creating final energy CSV: {fin_energy_csv_name}")
        fin_energy_df = df[cols + ["PJ"]]
        fin_energy_df = fin_energy_df.groupby(cols, as_index=False).sum()
        if not fin_energy_df.empty:
            fin_energy_df.to_csv(output_path / f"{fin_energy_csv_name}.csv", index=False)

@click.command()
@click.argument("input_path", type=click.Path(exists=True))
@click.option(
    "--no-en-int",
    is_flag=True,
    help="Skip processing of energy intensity files",
)
@click.option(
    "--en-int-only",
    is_flag=True,
    help="Only process energy intensity files (skip CSV creation)",
)
def process_msm22_files(input_path, no_en_int, en_int_only):
    """Process MSM22 files - creates CSVs and processes energy intensity by default.

    [bold green]Arguments:[/]

    [yellow]input_path[/]: Path to directory containing MSM22 Excel files

    [bold blue]Description:[/]

    This tool performs two operations:
    1. Converts all sheets from all MSM22 Excel files in the directory into combined CSV files
    2. Processes energy intensity (can be disabled with --no-en-int)

    Use --en-int-only to skip CSV creation and only process energy intensity files.
    """
    if en_int_only and no_en_int:
        raise click.UsageError("Cannot use both --en-int-only and --no-en-int")
    
    if not en_int_only:
        process_msm22_csvs(input_path)
    
    if not no_en_int or en_int_only:
        process_energy_intensity(input_path)

if __name__ == "__main__":
    process_msm22_files()

if __name__ == "__main__":
    process_msm22_files()  # Changed to use the CLI interface
